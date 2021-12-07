import sys
import yaml
import json
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.styles import Font

def create_xlsx(xlsxname):
    # ファイルは1つでいい
    wb = openpyxl.Workbook()
    wb.create_sheet(title="★5")
    wb.create_sheet(title="★4")
    wb.create_sheet(title="★3")
    wb.create_sheet(title="★2")
    wb.remove(wb.worksheets[0])
    return wb

def write_xlsx(ri, sheet, reach, moji, bukisyu, reach_C, buki_colors, borders, font):
    # 一旦ユニット名でソートかけましょうかね
    for bukikey in reach.keys():
        unitdata = reach[bukikey]
        reach[bukikey] = sorted(unitdata, key=lambda x: x[2], reverse=True)
    for bukikey in reach.keys():
        unitdata = reach[bukikey]
        for i in range(len(unitdata)):
            unit = unitdata[i]
            sheet.cell(row=ri, column=4, value=moji).fill = reach_C
            sheet.cell(row=ri, column=4).font = font
            sheet.cell(row=ri, column=5, value=bukisyu[bukikey-1]).fill = buki_colors[bukikey-1]
            sheet.cell(row=ri, column=5).font = font
            sheet.cell(row=ri, column=6, value=unit[2]).font = font
            sheet.cell(row=ri, column=3).border = borders[i == len(unitdata)-1]
            sheet.cell(row=ri, column=4).border = borders[i == len(unitdata)-1]
            sheet.cell(row=ri, column=5).border = borders[i == len(unitdata)-1]
            sheet.cell(row=ri, column=6).border = borders[i == len(unitdata)-1]
            ri += 1
    return ri


def main():
    with open('config.yaml', 'r',encoding="utf-8_sig") as yml:
        config = yaml.safe_load(yml)

    json_name = config["korin"]["unit_zokusei"]
    xlsxname = config["unit_syozi"]["xlsxname"]

    with open(json_name+'.json',encoding="utf-8_sig") as f:
        unit_ygarabo = json.load(f)

    print("unitdata: ", len(unit_ygarabo))


    # xlsxシート作成
    xlsx = create_xlsx(xlsxname)
    all_data = {"5":[], "4":[], "3":[], "2":[]}
    for unit in unit_ygarabo:
        unit = unit_ygarabo[unit]
        # ★順でソート
        # 後衛,中衛,前衛の順
        # その順の中でも武器種で分けられると所持状態を確認しやすい
        unitREACH = int(unit["reach"])
        unitRARITY = unit["rarity"]
        unitWEAPON = int(unit["weaponnum"])
        if unitWEAPON == 7:
            # 回復ユニットは使わない
            continue
        # 計算に必要なパラメータも抜き出しておく…？
        # いや、結局所持状態からyugarabo_jsonにマスクをかけて計算するからいらないかな…？
        unitNAME = unit["slug"]
        unitZINKUSI = unit["type"]
        all_data[unitRARITY].append([unitREACH,unitWEAPON,unitNAME,unitZINKUSI])

    # unitデータをとったので、ソートして分類していく
    fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFFF99', bgColor='FFFF99')
    border = Border(top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'), 
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000')
    )
    yoko_border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000')
    )
    ue_border = Border(bottom=Side(style='thin', color='000000'), 
                        left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000')
    )
    borders = [yoko_border, ue_border]

    kouei_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='BEEDCB', bgColor='BEEDCB')
    tyuei_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='BEE0ED', bgColor='BEE0ED')
    zenei_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFCCDA', bgColor='FFCCDA')
    zan_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFD699', bgColor='FFD699')
    totu_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='E8BADF', bgColor='E8BADF')
    da_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFF899', bgColor='FFF899')
    yumi_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFD699', bgColor='FFD699')
    mahou_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='E8BADF', bgColor='E8BADF')
    zyuu_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFF899', bgColor='FFF899')
    buki_colors = [zan_C, totu_C, da_C, yumi_C, mahou_C, zyuu_C]
    font = Font(name='ロックンロール One')

    bukisyu = ["斬撃","突撃","打撃","弓矢","魔法","銃撃"]
    for rea_index in range(4):
        rearity = 5-rea_index
        kouei = {1:[],2:[],3:[],4:[],5:[],6:[]}
        tyuei = {1:[],2:[],3:[],4:[],5:[],6:[]}
        zenei = {1:[],2:[],3:[],4:[],5:[],6:[]}
        for unit in all_data[str(rearity)]:
            if unit[0] <= 50:
                zenei[unit[1]].append(unit)
            elif 50 < unit[0] and unit[0] <= 150:
                tyuei[unit[1]].append(unit)
            elif 150 < unit[0]:
                kouei[unit[1]].append(unit)

        # ここから書き込み 後衛、中衛、前衛の順で
        sheet = xlsx['★'+str(rearity)]
        sheet.column_dimensions['F'].width = 40
        sheet.cell(row=5, column=3, value="所持状況").font = font
        sheet.cell(row=5, column=4, value="リーチ").font = font
        sheet.cell(row=5, column=5, value="武器種").font = font
        sheet.cell(row=5, column=6, value="名前").font = font
        for rows in sheet['C5':'F5']:
            for cell in rows:
                cell.fill = fill
                cell.border = border
        ri = 6
        ri = write_xlsx(ri, sheet, kouei, "後衛", bukisyu, kouei_C, buki_colors, borders, font)
        ri = write_xlsx(ri, sheet, tyuei, "中衛", bukisyu, tyuei_C, buki_colors, borders, font)
        ri = write_xlsx(ri, sheet, zenei, "前衛", bukisyu, zenei_C, buki_colors, borders, font)
    xlsx.save(xlsxname + '.xlsx')

# 4部位水打銃有利
# 5部位風斬魔有利

if __name__ == "__main__":
    main()