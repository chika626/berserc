from os import closerange
import math
import yaml
import openpyxl
from matplotlib import pyplot as plt
from tqdm import tqdm
import numpy as np
from operator import itemgetter
from openpyxl.styles import Border, Side, Font


def write_xlsx(ri, sheet, best_unit, bukisyu, reach_C, buki_colors, borders, font):
    # 一旦ユニット名でソートかけましょうかね
    for i in range(len(best_unit)):
        unit = best_unit[i]
        sheet.cell(row=ri, column=3, value='★'+str(unit["rearity"])).fill = reach_C
        sheet.cell(row=ri, column=3).font = font
        sheet.cell(row=ri, column=4, value=unit["reach"]).fill = reach_C
        sheet.cell(row=ri, column=4).font = font
        sheet.cell(row=ri, column=5, value=bukisyu[unit["weapon"]-1]).fill = buki_colors[unit["weapon"]-1]
        sheet.cell(row=ri, column=5).font = font
        sheet.cell(row=ri, column=6, value=unit["name"]).font = font
        sheet.cell(row=ri, column=7, value=', '.join(unit["rune"])).font = font
        sheet.cell(row=ri, column=8, value=', '.join(unit["assist"])).font = font
        sheet.cell(row=ri, column=9, value=str(unit["unitB"]*100)+'%').font = font
        sheet.cell(row=ri, column=10, value=unit["HPS"]).font = font
        sheet.cell(row=ri, column=11, value=unit["dps"]).font = font
        if "knight" in unit.keys():
            sheet.cell(row=ri, column=12, value=unit["knight"]).font = font
        if "riferune" in unit.keys():
            if unit["riferune"] != 0:
                sheet.cell(row=ri, column=13, value=unit["riferune"]).font = font
        if "guardrune" in unit.keys():
            if unit["guardrune"] != 0:
                sheet.cell(row=ri, column=14, value=unit["guardrune"]).font = font
        # sheet.cell(row=ri, column=3).border = borders[i == len(best_unit)-1]
        # sheet.cell(row=ri, column=4).border = borders[i == len(best_unit)-1]
        # sheet.cell(row=ri, column=5).border = borders[i == len(best_unit)-1]
        # sheet.cell(row=ri, column=6).border = borders[i == len(best_unit)-1]
        ri += 1
    '''
    sheet.cell(row=5, column=3, value="レア").font = font
    sheet.cell(row=5, column=4, value="リーチ").font = font
    sheet.cell(row=5, column=5, value="武器種").font = font
    sheet.cell(row=5, column=6, value="名前").font = font
    sheet.cell(row=5, column=7, value="ルーン").font = font
    sheet.cell(row=5, column=8, value="アシスト").font = font
    sheet.cell(row=5, column=9, value="防護力").font = font
    sheet.cell(row=5, column=10, value="HP").font = font
    sheet.cell(row=5, column=11, value="DPS").font = font
    '''

def get_highdpsmaster(fulldata, config):
    knight = 0.15001
    knight_step = 0.005
    m = 37
    # 取るのは上から24人
    pickup_unit = int(config["unit_syozi"]["pickup"])
    hd_x_unitdata = {}
    for data in fulldata:
        hd_x_unitdata[data["hidame"]] = data["unit_datas"]
    # 被ダメ毎に見ていく
    # 見ていくうちにナイトで下側も計算する
    # dpsでソートする必要があるので、配列が必要

    start_reach = 15
    end_reach = 150

    zenei_limit = 9

    best_hid = 0
    best_adps = 0
    best_unit = []

    for hidamage in tqdm(hd_x_unitdata.keys()):
        unit_datas = hd_x_unitdata[hidamage]
        # ナイトの方が上回った時にswapをする
        for knight_x in range(m):
            # ナイトを使う場合、全て固定しないといけない
            knight_M = knight + knight_step * knight_x
            Khid = hidamage / (1-knight_M*2)
            # ダメージは大きくなるので、ラインは100区切りの下側で見る
            # ラインが存在しない場合、ナイトは使えない
            Khid = int(math.floor(Khid/100) * 100)
            use_Knight = Khid in hd_x_unitdata.keys()
            # knight設定し始めるリーチを決める
            for reach_x in range(16):
                reach = reach_x * 5 + start_reach
                temp_datas = unit_datas.copy()
                for i, unit in enumerate(temp_datas):
                    if use_Knight:
                        k_unit = hd_x_unitdata[Khid][i]
                        # 中衛じゃないと意味ない
                        if reach >= unit["reach"]:
                            # swapする
                            temp_datas[i] = k_unit.copy()
                            temp_datas[i]["knight"] = knight_M
                temp_datas = sorted(temp_datas, key=itemgetter('dps'), reverse=True)
                
                # 上から24人とswapする
                adps = 0
                tyuei_select = 0
                # まずは中衛の上位から処理する
                tyuei_sel_datas = []
                for x in temp_datas:
                    if x['rear guard']:
                        tyuei_sel_datas.append(x)
                    else:
                        if tyuei_select < zenei_limit:
                            tyuei_sel_datas.append(x)
                            tyuei_select += 1
                temp_datas = tyuei_sel_datas
                for i in range(pickup_unit):
                    adps += temp_datas[i]["dps"]
                if adps > best_adps:
                    best_adps = adps
                    best_hid = hidamage
                    best_unit = temp_datas[:pickup_unit]
    print(best_adps, " ", best_hid)
    for unit in best_unit:
        print(unit)
    # 出力もするけど保存もしたい

    fill = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFFF99', bgColor='FFFF99')
    border = Border(top=Side(style='thin', color='000000'), 
                    bottom=Side(style='thin', color='000000'), 
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000')
    )
    yoko_border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000')
    )
    ue_border = Border(bottom=Side(style='thin', color='000000'), 
                        left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000')
    )
    borders = [yoko_border, ue_border]

    kouei_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='BEEDCB', bgColor='BEEDCB')
    tyuei_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='BEE0ED', bgColor='BEE0ED')
    zenei_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFCCDA', bgColor='FFCCDA')
    zan_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFD699', bgColor='FFD699')
    totu_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='E8BADF', bgColor='E8BADF')
    da_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFF899', bgColor='FFF899')
    yumi_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFD699', bgColor='FFD699')
    mahou_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='E8BADF', bgColor='E8BADF')
    zyuu_C = openpyxl.styles.PatternFill(patternType='solid',fgColor='FFF899', bgColor='FFF899')
    buki_colors = [zan_C, totu_C, da_C, yumi_C, mahou_C, zyuu_C]
    font = Font(name='ロックンロール One')

    bukisyu = ["斬撃","突撃","打撃","弓矢","魔法","銃撃"]
    xlsx = openpyxl.Workbook()
    xlsxname = config["unit_syozi"]["simulation_data"] + "ans"
    sheet = xlsx["Sheet"]
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 40
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 12
    sheet.column_dimensions['M'].width = 12
    sheet.column_dimensions['N'].width = 12
    sheet.cell(row=5, column=3, value="レア").font = font
    sheet.cell(row=5, column=4, value="リーチ").font = font
    sheet.cell(row=5, column=5, value="武器種").font = font
    sheet.cell(row=5, column=6, value="名前").font = font
    sheet.cell(row=5, column=7, value="ルーン").font = font
    sheet.cell(row=5, column=8, value="アシスト").font = font
    sheet.cell(row=5, column=9, value="防護力").font = font
    sheet.cell(row=5, column=10, value="HP").font = font
    sheet.cell(row=5, column=11, value="DPS").font = font
    sheet.cell(row=5, column=12, value="ナイト保持").font = font
    sheet.cell(row=5, column=13, value="ライフ保持").font = font
    sheet.cell(row=5, column=14, value="ガード保持").font = font
    

    for rows in sheet['C5':'N5']:
        for cell in rows:
            cell.fill = fill
            cell.border = border
    ri = 6
    ri = write_xlsx(ri, sheet, best_unit, bukisyu, kouei_C, buki_colors, borders, font)
    xlsx.save(xlsxname + '.xlsx')
            




def main():
    with open('config/config.yaml', 'r' ,encoding="utf-8_sig") as yml:
        config = yaml.load(yml)
    fulldata = np.load(config["unit_syozi"]["simulation_data"] + ".npy" ,allow_pickle=True)
    get_highdpsmaster(fulldata, config)

    return
    # データを分解し、被ダメ当たりの総DPSを計算してみる
    adps = 0
    hid = 0
    for data in tqdm(fulldata):
        dps = 0
        for unitdata in data["unit_datas"]:
            print(unitdata["dps"])
            dps += int(unitdata["dps"])
        if dps > adps:
            adps = dps
            hid = data["hidame"]
    for data in tqdm(fulldata):
        if hid != data["hidame"]:
            continue
        for unitdata in data["unit_datas"]:
            print(unitdata)




    # 4部位水打銃有利
    # 5部位風斬魔有利

if __name__ == "__main__":
    main()