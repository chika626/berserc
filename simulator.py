import enum
from os import closerange
import sys
# from numpy.core.fromnumeric import argmax
import yaml
import json
import openpyxl
import itertools
from matplotlib import pyplot as plt
from tqdm import tqdm
import numpy as np


# アシスト、防護力の値、ルーンのパターン保持し、DPSを計算するclass
class PATTERN:
    def __init__(self, assist=["135","",""], unitB = 0.03, rune = [""], rearity = 0):
        self.assist = assist
        self.unitB = unitB/100.0
        self.rune = rune
        self.rife_M = 0
        self.guard_M = 0
        self.dps = 0
        self.HPS = 0
        self.seven = False
        if rearity == 5 and assist[1] == "zoka":
            self.seven = True

    def get_rune_atk(self):
        # atkは33.000固定
        # 5%強化ならそう返す
        x = 1
        if self.rune.index("atk") % 2 == 0:
            x = 1.05
            if self.seven:
                x = 1.07
        return 0.33 * x

    def get_rune_quick(self):
        # quickは33.000固定かつ5%強化
        x = 1.05
        if self.seven:
            x = 1.07
        return 0.33000 * x

    def get_rune_berse(self):
        # berseは33.000固定かつ5%強化
        x = 1.05
        if self.seven:
            x = 1.07
        return 0.33650 * x

    def get_rune_zokusei(self):
        # zokuseiは30.000固定
        if not "zokusei" in self.rune:
            return 0
        x = 1
        if self.seven:
            x = 1.02
        return 0.33 * x

    def get_guts(self):
        # ガッツは27.001固定なので30になる
        if not "guts" in self.rune:
            return 28
        return 30

    def get_rune_arch(self):
        # archは30.000固定
        if not "arch" in self.rune:
            return 0
        x = 1
        if self.seven:
            x = 1.02
        return 0.30 * x

    def is_get_rune_rife(self):
        # rifeは30.000固定
        if not "rife" in self.rune:
            return False
        return True
    def set_rune_rife(self, rife_M):
        self.rife_M = rife_M
        if self.seven:
            self.rife_M /= 1.02
    def get_rune_rife(self):
        return self.rife_M

    def is_get_rune_guard(self):
        # guardは30.000固定
        if not "guard" in self.rune:
            return False
        return True
    def set_rune_guard(self, guard_M):
        self.guard_M = guard_M
        if self.seven:
            self.guard_M /= 1.02
    def get_rune_guard(self):
        return self.guard_M

    def get_unitB(self):
        return self.unitB
    def get_as_yuri(self):
        if "yuri" in self.assist:
            return 2
        return 0

class UNIT:
    def __init__(self, argument):
        # ユニット攻撃力
        self.ATK = argument["atk"]
        # ユニット攻撃間隔
        self.INTER = argument["inter"]
        # ユニット攻撃体数
        self.ATK_V = argument["atk_b"]
        # ユニットの属性補正
        self.HOSEI = argument["hosei"]
        # マインドのおまけ効果(9 or 12)
        self.MIND_OMAKE = argument["omake"]
        # 特攻かどうかのフラグ
        self.SPECIAL = argument["flagspecial"]
        # ユニットガッツ(出撃時28なので28で初期化)
        # いまのあれが12なので12で初期化
        self.GUTS = 28
        # ユニットアシスト(耐久指数で変化するので0で初期化)
        self.ASSIST = 0
        # ルーン枠の数
        self.LIM_RUNE = argument["rearity"] + 1
        self.REARITY = argument["rearity"]
        # リーチ(上位ルーンの装備が出来るのかどうなのか)
        self.can_zyoui = argument["can_zyoui"]
        # 名前と武器種類
        self.name = argument["name"]
        self.weapon = argument["weapon"]
        self.HP = argument["hp"]
        self.reach = argument["reach"]


    # 救援のアシスト、ルーン、防護力のパターン生成場所 
    def create_pattern(self):
        # アシスト、2パターン
        assists = [["135","yuri","yuri"],["135","zoka","zoka"]]
        # unitBs = [0.03, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
        unitBs = [0.03]
        unitBs_start = 0.5
        unitBs_step = 0.25
        unitBs_end = 13.0
        for i in range(int((unitBs_end-unitBs_start)/unitBs_step)):
            unitBs.append(unitBs_start+i*unitBs_step)
        runes = ["zokusei","guts","rife","guard",""]
        runes_a = ["zokusei","guts","rife","guard","arch"]
        ret = []
        for assist in assists:
            for unitB in unitBs:
                # アシストと防護力が決定したので残りルーン枠がわかる
                # アタクイベルセの3枠を引く
                nokoriWAKU = self.LIM_RUNE - 3
                # 増加が入ってるなら+1する
                if assist[1] == "zoka":
                    nokoriWAKU = min(3, nokoriWAKU + 1)
                # runes or runes_a C nokoriWAKUで計算する
                set_N = runes
                # 武器種が弓ならアーチェも考える
                if self.weapon == 4:
                    set_N = runes_a
                for rune_p in itertools.combinations(set_N, nokoriWAKU):
                    # ルーンも決まったのでPATTERNとして保存する
                    # ★5 or ★4増加
                    if (self.LIM_RUNE - 1 == 5) or (self.LIM_RUNE - 1 == 4 and assist[1] == "zoka"):
                        rune = ["berse",rune_p[0],"quick",rune_p[1],"atk",rune_p[2]]
                    # ★4 or ★3増加
                    elif (self.LIM_RUNE - 1 == 4) or (self.LIM_RUNE - 1 == 3 and assist[1] == "zoka"):
                        rune = ["berse",rune_p[0],"quick",rune_p[1],"atk"]
                    # ★3
                    elif (self.LIM_RUNE - 1 == 3):
                        rune = ["berse","atk","quick",rune_p[0]]
                    ret.append([assist,unitB,rune])
        
        return ret

    def dps_calc(self, pt, hiDamage, seedBREAK, monsterV, runeICHRYS = 0, runeENHANCER = 0, seedSOUL = 0):
        # いつものようにdpsを計算しますが、いろいろ欠けてない…？
        calcATK = cuest_ATK(self.ATK, pt.get_guts(), pt.get_rune_atk(), 0, pt.get_unitB())
        calZHOSEI = cuest_HOSEI(self.HOSEI, pt.get_as_yuri(), self.MIND_OMAKE, pt.get_rune_zokusei(), seedBREAK, runeICHRYS, runeENHANCER)
        damage = YATK(calcATK, calZHOSEI, seedSOUL, self.SPECIAL)
        # ライフ ガードがある場合は魔力値を設定してあげないといけない
        if pt.is_get_rune_rife():
            rife_M = 0.15001
            for x in range(25):
                rife_M = 0.15001 + x * 0.005
                cuest_HP = HPS(self.HP, pt.get_guts(), 0, pt.get_unitB(), rife_M, 0)
                # 5%以上耐えられるようになったら終わり
                if cuest_HP > (hiDamage*1.05):
                    break
            # もしだめでも27.001は持たせる
            pt.set_rune_rife(rife_M)
            # print("     rife: ", rife_M)
        # ガードでも同じことをする
        if pt.is_get_rune_guard():
            guard_M = 0.15001
            for x in range(25):
                guard_M = 0.15001 + x * 0.005
                cuest_HP = HPS(self.HP, pt.get_guts(), 0, pt.get_unitB(), pt.get_rune_rife(), guard_M)
                # 5%以上耐えられるようになったら終わり
                if cuest_HP > (hiDamage*1.05):
                    break
            pt.set_rune_guard(guard_M)
            # print("     guard: ", guard_M)

        # 死なない設定終わったらHP計算する
        cuest_HP = HPS(self.HP, pt.get_guts(), 0, pt.get_unitB(), pt.get_rune_rife(), pt.get_rune_guard())
        pt.HPS = cuest_HP

        # HP足りてなさそうなら計算を打ち切って0を返す
        if hiDamage > cuest_HP:
            pt.dps = 0
            return 0

        percentage = 1 - (cuest_HP - hiDamage) / cuest_HP
        reDPS = DPS(damage, self.INTER, pt.get_rune_quick(), pt.get_guts(), self.ATK_V, monsterV, percentage, pt.get_rune_berse(), pt.get_rune_arch())
        pt.dps = reDPS
        return reDPS

    def Berserk_Adjustment(self, hiDamage):
        # まずはパターンの作成をする
        self.pattern = []
        for assist, unitB, rune in self.create_pattern():
            pt = PATTERN(assist, unitB, rune, (self.LIM_RUNE-1))
            self.pattern.append(pt)

        # パターン毎にDPSを計算する 耐えられないなら0にする
        max_dps = 0
        ret = self.pattern[0]
        for pt in self.pattern:
            dps = self.dps_calc(pt, hiDamage, seedBREAK=0.315, monsterV=3)
            if dps > max_dps:
                ret = pt
                max_dps = dps
            # print("as: ",pt.assist,"  B: ", pt.unitB, "  rune: ",pt.rune , "  dps: ", dps , " HPS: ",pt.HPS)
        return ret

    # def HPcalc_and_safeLINE(self, line, seedBREAK, runeICHRYS, runeENHANCER, seedSOUL, monsterASS):
    #     # lineを超えるような体力にあるか、ないならどうすればいいのかを検討する
    #     # まずは防護力0.03、上位なし、ガッツなしで耐えられるのか確認
    #     cuestHP = HPS(self.HP, self.GUTS, runeHIGH=0, unitB=0.03, runeRIFE=0)
    #     if cuestHP > line:
    #         # ラインを超えてる場合 防護力は0.03で確定 上位も使わない
    #         self.ASSIST = 0.03
    #         # ★5の場合
    #         if self.LIM_RUNE-1 == 5:
    #             self.runeDEPLOY = ["atk","guts","quick","zokusei","berse",""]
    #             self.runeMVALUE = [33.00, 27.001, 33.00, 30.00, 33.00, None]
    #             # ★5はアーチェを入れる余裕がある
    #             if self.weapon == 4:
    #                 self.runeDEPLOY[5] = "arche"
    #                 self.runeMVALUE[5] = 30.00
    #         # ★4の場合
    #         elif self.LIM_RUNE-1 == 4:
    #             self.runeDEPLOY = ["atk","guts","quick","zokusei","berse"]
    #             self.runeMVALUE = [33.00, 27.001, 33.00, 30.00, 33.00]
    #             # ★4はアーチェを入れるならGutsを抜く
    #             if self.weapon == 4:
    #                 self.runeDEPLOY[1] = "arche"
    #                 self.runeMVALUE[1] = 30.00
    #         # ★3の場合
    #         elif self.LIM_RUNE-1 == 4:
    #             # 属性をいれられて耐えるならそれでいい 優先度は属性 > arche > 有利2
    #             self.runeDEPLOY = ["berse","atk","quick","zokusei"]
    #             self.runeMVALUE = [33.00, 33.00, 33.00, 30.00]
    #     else:
    #         # 耐えない場合 どうすんのお前
    #         # 優先度 ガッツ＞上位＞防護力＞他
    #         pass

    # def calcrate_assist(self, line,unit,rune,assist,seed,mons):
    #     c_ATK = cuest_ATK(unit["atk"], unit["gtus"], rune["atk"], rune["high"], unit["assist"])
    #     zokuseiHOSEI = cuest_HOSEI(unit["hosei"], assist["yuri"], unit["mind_omk"], rune["zokusei"], seed["break"], rune["ex"], rune["enha"])
    #     damage = YATK(c_ATK, zokuseiHOSEI, seed["soul"], unit["special"])
    #     cuest_HP = HPS(unit["hp"], unit["gtus"], rune["high"], unit["assist"], rune["rife"])
    #     percentage = 1 - line / cuest_HP
    #     reDPS = DPS(damage, unit["inter"], rune["quick"], unit["guts"], unit["assist"], mons, percentage, rune["berse"], rune["arche"])
    #     return reDPS
    



def cuest_ATK(unitATK, unitGUTS, runeATK, runeHIGH, unitB):
    return unitATK*((9+unitGUTS)/10)*(1+runeATK+runeHIGH/5+unitB)

def mindYUURIHOSEI(unitHOSEI , selectnum):
    return (unitHOSEI*0.06)*selectnum

def cuest_HOSEI(unitHOSEI, mindYURI, mindOMAKE, runeZOKUSEI, seedBREAK, runeICHRYS, runeENHANCER):
    return ((unitHOSEI + mindYUURIHOSEI(unitHOSEI, mindYURI) + mindOMAKE) * (1+runeZOKUSEI)) + (seedBREAK * 5) + (runeICHRYS / 3) + (runeENHANCER / 3)

def YATK(cuest_ATK, cuest_HOSEI, seedSOUL, flagSPECIAL):
    return cuest_ATK * cuest_HOSEI * (1+seedSOUL)**5 * (1+flagSPECIAL*.35)

def DPS(damage, unitKANAKU, runeQUICK, unitGUTS, unitASS, monsterASS, wariai, runeBERSE, runeARCHE = 0):
    unitGUTS = min(11, unitGUTS)
    all_damage = damage * min(unitASS, monsterASS) / (unitKANAKU * (1 - runeQUICK) * (1 - (unitGUTS - 1)* 0.04))
    return  all_damage + (all_damage * (wariai) * (runeBERSE) * (runeBERSE) * 8) + (runeARCHE * all_damage * 0.5)

def HPS(unitHP, unitGUTS, runeHIGH, unitB, runeRIFE = 0, runeGURAD = 0):
    return unitHP * (0.9 + unitGUTS/10) * (1 + runeRIFE + runeHIGH/2 + unitB) / (1-unitB-runeGURAD)

def get_syozi_unitmask(xlsxname):
    wb = openpyxl.load_workbook(xlsxname+".xlsx")
    syozi_mask = {}
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        syozi_column = 3
        uname_column = 6
        ri = 6
        while True:
            # ユニット名が空だったときに落とす
            syozi = sheet.cell(row=ri, column=syozi_column)
            unitname = sheet.cell(row=ri, column=uname_column)
            if unitname.value is None:
                break
            syozi_mask[unitname.value] = (not syozi.value is None)
            ri += 1
    return syozi_mask

def tokkou(bukisyu, s0, s1, unitbuki):
    return (bukisyu.index(s0)+1) == unitbuki or (bukisyu.index(s1)+1) == unitbuki

def main():
    # configを読む
    with open('config/config.yaml', 'r' ,encoding="utf-8_sig") as yml:
        config = yaml.load(yml)

    # 所持状況表を読み取る
    syozi_mask = get_syozi_unitmask(config["unit_syozi"]["xlsxname"])
    
    # ユニットの所持情報から所持ユニットのみの情報にする
    # UNITを保持し、全部にBerserk_Adjustmentさせて結果を返させる
    with open(config["korin"]["unit_zokusei"]+'.json',encoding="utf-8_sig") as f:
        unit_ygarabo = json.load(f)
    MYUNITS = []
    monster_zonkusei = config["korin"]["monster_zokusei"]
    omake_dict = {"cfire":0.09, "cwater":0.09 , "cwind":0.09, "clight":0.12, "cdark":0.12}
    bukisyu = ["斬撃","突撃","打撃","弓矢","魔法","銃撃"]

    for unit in unit_ygarabo:
        unit = unit_ygarabo[unit]
        unitNAME = unit["slug"]
        unitWEAPON = int(unit["weaponnum"])
        if unitWEAPON == 7 or not syozi_mask[unitNAME]:
            # 回復と持ってないユニットは読み飛ばす
            continue
        atk = int(unit["atkmax"])
        inter = float(unit["inter"])
        atk_b = int(unit["assault"])
        hosei = float(unit[monster_zonkusei])/100.0
        omake = omake_dict[monster_zonkusei]
        flagspecial = tokkou(bukisyu, config["korin"]["s0"], config["korin"]["s1"], int(unit["weaponnum"]))
        rearity = int(unit["rarity"])
        can_zyoui = (int(unit["reach"]) > 150)
        hp = int(unit["hpmax"])
        argument = {"atk":atk,
                    "inter":inter,
                    "atk_b":atk_b, 
                    "hosei":hosei, 
                    "omake":omake, 
                    "flagspecial":flagspecial, 
                    "rearity":rearity,
                    "reach": int(unit["reach"]),
                    "can_zyoui":can_zyoui,
                    "name":unitNAME,
                    "weapon":unitWEAPON,
                    "hp":hp,
        }
        unitdata = UNIT(argument)
        MYUNITS.append(unitdata)
    print(len(MYUNITS))

    # # サロ
    # debug_unit = MYUNITS[6]
    # # ロッシェ
    # debug_unit = MYUNITS[12]
    # # ナノン
    # debug_unit = MYUNITS[26]
    # # パユ
    # debug_unit = MYUNITS[47]
    # # カナカエイア
    # debug_unit = MYUNITS[62]
    # ルジャS
    # debug_unit_2 = MYUNITS[122]

    # debug_unit = MYUNITS[1]

    # # ここから計算をはじめるフェーズ
    # print(debug_unit.name)
    # pt = debug_unit.Berserk_Adjustment(3000)
    # step = 100
    # x_line = []
    # y_dps = []
    # y_dps_2 = []
    # for x in range(500):
    #     hidamage = step*x
    #     pt = debug_unit.Berserk_Adjustment(hidamage)
    #     pt_2 = debug_unit_2.Berserk_Adjustment(hidamage)
    #     if pt is None:
    #         break
    #     x_line.append(hidamage)
    #     y_dps.append(pt.dps)
    #     y_dps_2.append(pt_2.dps)
    #     # print("hid: ", hidamage, "as: ",pt.assist,"  B: ", pt.unitB, "  rune: ",pt.rune , "  dps: ", pt.dps , " HPS: ",pt.HPS)
    # # plt.title(debug_unit.name, fontname="RocknRoll One")
    # plt.plot(x_line, y_dps)
    # plt.plot(x_line, y_dps_2)
    # plt.legend(["ナジュム","ルジャS"], prop={"family":"RocknRoll One"})
    # plt.xlabel("耐久指数", fontname="RocknRoll One")
    # plt.ylabel("DPS", fontname="RocknRoll One")
    # # plt.legend()
    # plt.show()

    # return

    # 最低ラインから+100くらいで計算、付けられるルーン、DPSを考慮した
    min_line = 20000
    step = 1000
    max_line = 40000

    m = int((max_line-min_line)/step)
    store_data = []
    for x in tqdm(range(m)):
        now_line = min_line + step * x
        hd_data = {"hidame":now_line, "unit_datas":[] }
        
        # ラインでベルセを含めたDPSを計算
        for unit in MYUNITS:
            # パターンが帰る
            pt = unit.Berserk_Adjustment(now_line)
            # 保持したいのは
            # ユニット判断のためにレアリティ、武器種、名前
            # 総DPS、ルーン、アシスト
            unitDICT = {"rearity": unit.REARITY,
                        "weapon": unit.weapon,
                        "name": unit.name,
                        "dps": pt.dps,
                        "rune": pt.rune,
                        "assist": pt.assist,
                        "unitB": pt.unitB, 
                        "riferune": pt.rife_M,
                        "guardrune": pt.guard_M,
                        "HPS": pt.HPS,
                        "rear guard": unit.can_zyoui,
                        "reach":unit.reach ,
            }
            hd_data["unit_datas"].append(unitDICT)
        store_data.append(hd_data)
    store_data = np.array(store_data)
    np.save(config["unit_syozi"]["simulation_data"], store_data)

    # 4部位水打銃有利
    # 5部位風斬魔有利

if __name__ == "__main__":
    main()